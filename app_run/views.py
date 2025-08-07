from decimal import Decimal, ROUND_HALF_UP

from django.conf import settings
from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q, Avg, Max
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from geopy.distance import geodesic
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet

from app_run.distance import calculate_distance
from app_run.models import Run, AthleteInfo, Challenge, Position, CollectibleItem, Subscribe
from app_run.serializers import RunSerializer, UserSerializer, AthleteInfoSerializer, ChallengeSerializer, \
    PositionSerializer, CollectibleItemSerializer, AthleteDetailSerializer, CoachDetailSerializer, \
    AthleteChallengeSerializer, SubscribeSerializer


class RunUserPagination(PageNumberPagination):
    page_size_query_param = 'size'


class AnalyticsForCoachView(APIView):
    def get(self, request, *args, **kwargs):
        coach = User.objects.filter(id=self.kwargs.get('id'), is_superuser=False).first()
        if not coach:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if not coach.is_staff:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        analytics = Run.objects.filter(athlete__coaches_coach=coach.id).values('athlete').annotate(
            avg_speed=Avg('speed'),
            total_distance=Sum('distance'),
            max_distance=Max('distance'))

        total_run_value, total_run_user = 0, 0
        speed_avg_value, speed_avg_user = 0, 0
        longest_run_user, longest_run_value = 0, 0

        for athlete in analytics:
            if athlete['max_distance'] > longest_run_value:
                longest_run_value = athlete['max_distance']
                longest_run_user = athlete['athlete']

            if athlete['avg_speed'] > speed_avg_value:
                speed_avg_value = athlete['avg_speed']
                speed_avg_user = athlete['athlete']

            if athlete['total_distance'] > total_run_value:
                total_run_value = athlete['total_distance']
                total_run_user = athlete['athlete']

        return Response({
            'longest_run_user': longest_run_user,

            'longest_run_value': longest_run_value,

            'total_run_user': total_run_user,

            'total_run_value': total_run_value,

            'speed_avg_user': speed_avg_user,

            'speed_avg_value': speed_avg_value
        }, status=status.HTTP_200_OK)


class RateCoachView(APIView):
    def post(self, request, *args, **kwargs):
        athlete = User.objects.filter(id=request.data.get('athlete'), is_staff=False, is_superuser=False).first()
        if not athlete:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        coach = User.objects.filter(id=self.kwargs.get('id'), is_superuser=False).first()
        if not coach:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if not coach.is_staff:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        subscription = Subscribe.objects.filter(athlete=athlete, coach=coach).first()
        if subscription:
            if 'rating' not in request.data:
                return Response({'error': 'Поле rating обязательно'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                rating = int(request.data.get('rating'))
            except (ValueError, TypeError):
                return Response({'error': 'Рейтинг должен быть числом'}, status=status.HTTP_400_BAD_REQUEST)
            serializer = SubscribeSerializer(data={'coach': coach.id, 'athlete': athlete.id, 'rating': rating})
            if serializer.is_valid():
                subscription.rating = rating
                subscription.save()
                return Response({'rating': subscription.rating}, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        return Response({'error': 'Пользователь не подписан на этого тренера'}, status=status.HTTP_400_BAD_REQUEST)


class ChallengeSummaryView(APIView):
    def get(self, request, *args, **kwargs):
        challenges = Challenge.objects.select_related('athlete').all().order_by('full_name')
        challenge_athletes = {}
        for challenge in challenges:
            if challenge.full_name not in challenge_athletes:
                challenge_athletes[challenge.full_name] = []
            challenge_athletes[challenge.full_name].append(challenge.athlete)

        answer = []

        for challenge, athletes in challenge_athletes.items():
            data = {
                'name_to_display': challenge,
                'athletes': AthleteChallengeSerializer(athletes, many=True).data
            }
            answer.append(data)

        return Response(status=status.HTTP_200_OK, data=answer)


class SubscribeView(APIView):
    def post(self, request, *args, **kwargs):
        athlete = User.objects.filter(id=request.data.get('athlete'), is_staff=False, is_superuser=False).first()
        if not athlete:
            return Response(status=status.HTTP_400_BAD_REQUEST)
        coach = User.objects.filter(id=self.kwargs.get('id'), is_superuser=False).first()
        if not coach:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if not coach.is_staff:
            return Response(status=status.HTTP_400_BAD_REQUEST)

        if Subscribe.objects.filter(athlete=athlete, coach=coach).exists():
            return Response(status=status.HTTP_400_BAD_REQUEST)
        Subscribe.objects.create(athlete=athlete, coach=coach)
        return Response(status=status.HTTP_200_OK)


class RunViewSet(ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    pagination_class = RunUserPagination


class PositionViewSet(ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['run']

    def perform_create(self, serializer):
        run = serializer.validated_data['run']
        positions = Position.objects.filter(run=run).order_by('date_time')
        if positions.exists():
            previous_position = positions.last()
            distance_previous_to_current = geodesic([previous_position.latitude, previous_position.longitude],
                                                    [serializer.validated_data['latitude'],
                                                     serializer.validated_data['longitude']]).m
            previous_time = previous_position.date_time
            last_time = serializer.validated_data['date_time']
            distance = previous_position.distance * 1000 + Decimal(str(distance_previous_to_current))
            time = (last_time - previous_time).total_seconds()
            if time > 0 and distance_previous_to_current > 0:
                speed = Decimal(str(distance_previous_to_current / time))
            else:
                speed = Decimal(previous_position.speed) if previous_position.speed else Decimal('0.00')

            if distance == 0:
                speed = Decimal('0.00')
        else:
            distance, speed = Decimal('0.0000'), Decimal('0.00')
        serializer.save(distance=(distance / 1000).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP),
                        speed=speed.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


class UserViewSet(ReadOnlyModelViewSet):
    queryset = User.objects.all().annotate(runs_finished=Count('run', filter=Q(run__status='finished')))
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']
    pagination_class = RunUserPagination

    def get_serializer_class(self):
        if self.action == 'list':
            return UserSerializer
        elif self.action == 'retrieve':
            user = User.objects.filter(id=self.kwargs.get('pk'), is_staff=False).first()
            return AthleteDetailSerializer if user else CoachDetailSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        qs = self.queryset
        type = self.request.query_params.get('type', None)
        if self.action == 'list':
            if type == 'coach':
                qs = qs.filter(is_staff=1).annotate(rating=Avg('athletes__rating'))
            elif type == 'athlete':
                qs = qs.filter(is_staff=0)
            return qs.filter(is_superuser=False)
        elif self.action == 'retrieve':
            user_id = self.kwargs.get('pk')
            user = get_object_or_404(User, id=user_id)
            if user.is_staff:
                qs = qs.prefetch_related('athletes__athlete').annotate(rating=Avg('athletes__rating'))
            else:
                qs = qs.prefetch_related('coaches__coach')
        return qs.filter(is_superuser=False)


class RunStartView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'init':
            run.status = 'in_progress'
            run.save()
            return Response(status=status.HTTP_200_OK, data={'message': 'Забег начат'})
        return Response(status=status.HTTP_400_BAD_REQUEST)


class RunStopView(APIView):
    def post(self, request, run_id):
        run = get_object_or_404(Run, id=run_id)
        if run.status == 'in_progress':

            # --------------- Calculate distance and number of finished runs ---------------

            run.status = 'finished'
            run.distance = calculate_distance(run)
            run.save()
            finished_runs = Run.objects.filter(athlete=run.athlete, status='finished')
            finished_runs_data = finished_runs.aggregate(Count('id'), Sum('distance'))
            finished_runs_data['distance__sum'] = finished_runs_data['distance__sum'] or 0

            # ------------------------- Calculate time and speed -------------------------

            positions = Position.objects.filter(run=run).order_by('date_time')
            if positions.count() >= 2:
                min_date, max_date = positions.first().date_time, positions.last().date_time
                time = (max_date - min_date).total_seconds()
                if time > 0:
                    run.run_time_seconds = int(time)
                    avg_speed = positions.aggregate(avg_speed=Avg('speed'))['avg_speed']
                    run.speed = round(avg_speed, 2) if avg_speed else 0
                else:
                    run.run_time_seconds, run.speed = 0, 0
            else:
                run.run_time_seconds, run.speed = 0, 0

            run.save()

            # ------------------------- Check Challenges -------------------------

            if finished_runs_data.get('id__count') == 10 and not Challenge.objects.filter(
                    athlete=run.athlete,
                    full_name=Challenge.CHALLENGE_10_RUNS).exists():
                Challenge.objects.create(full_name=Challenge.CHALLENGE_10_RUNS, athlete=run.athlete)

            if finished_runs_data.get('distance__sum') >= 50 and not Challenge.objects.filter(
                    athlete=run.athlete,
                    full_name=Challenge.CHALLENGE_50KM).exists():
                Challenge.objects.create(full_name=Challenge.CHALLENGE_50KM, athlete=run.athlete)

            if (
                    positions.count() >= 2 and run.run_time_seconds <= 600 and run.distance >= 2) and not Challenge.objects.filter(
                athlete=run.athlete,
                full_name=Challenge.CHALLENGE_2KM_10MIN).exists():
                Challenge.objects.create(full_name=Challenge.CHALLENGE_2KM_10MIN, athlete=run.athlete)

            return Response(RunSerializer(run).data, status=status.HTTP_200_OK)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class AthleteInfoView(APIView):
    def put(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        serializer = AthleteInfoSerializer(data=request.data)
        if serializer.is_valid():
            AthleteInfo.objects.update_or_create(user_id=user, defaults=serializer.validated_data)
            return Response(status=status.HTTP_201_CREATED, data={'message': 'Данные успешно добавлены'})
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request, user_id):
        user = get_object_or_404(User, id=user_id)
        info, created = AthleteInfo.objects.get_or_create(user_id=user, defaults={'weight': 0, 'goals': ''})
        serializer_data = AthleteInfoSerializer(info).data
        return Response(status=status.HTTP_200_OK, data=serializer_data)


@api_view(['GET'])
def show_collectible_items(request):
    return Response(CollectibleItemSerializer(CollectibleItem.objects.all(), many=True).data)


@api_view(['POST'])
def upload_collectible_items(request):
    file = request.FILES.get('file')
    if not file:
        return Response(status=status.HTTP_400_BAD_REQUEST, data={'error': 'Файл не передан'})

    xl = load_workbook(file)
    active_list = xl.active

    invalid_rows = []

    for row in active_list.iter_rows(min_row=2):
        row_data = {
            'name': row[0].value,
            'uid': row[1].value,
            'value': row[2].value,
            'latitude': row[3].value,
            'longitude': row[4].value,
            'picture': row[5].value
        }

        serializer = CollectibleItemSerializer(data=row_data)
        if serializer.is_valid():
            serializer.save()
        else:
            invalid_rows.append([i.value for i in row])

    return Response(status=status.HTTP_200_OK, data=invalid_rows)


@api_view(['GET'])
def show_challenges(request):
    challenges = Challenge.objects.all()
    athlete = request.GET.get('athlete')
    if athlete:
        challenges = challenges.filter(athlete__id=int(athlete))
    serializer_data = ChallengeSerializer(challenges, many=True).data
    return Response(serializer_data)


@api_view(['GET'])
def company_details_view(request):
    company_name = settings.COMPANY_NAME
    slogan = settings.SLOGAN
    contacts = settings.CONTACTS
    return Response({
        'company_name': company_name,
        'slogan': slogan,
        'contacts': contacts
    })
